import openpyxl

def find_last_data_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0

# Load the workbook and active sheet
excel = openpyxl.load_workbook("C:\\Users\\MathijesiJohnbritto\\Documents\\read.xlsx")
sheet = excel.active

# New data to add
data = [
    ["Alice", "Smith", 30, "Female", "1995-06-15"],
    ["Bob", "Johnson", 25, "Male", "2000-09-21"],
    ["Charlie", "Brown", 28, "Male", "1997-03-10"]
]

# Find the actual last row with data
last_row = find_last_data_row(sheet)

# Write new data starting after the last data row
for row_index, row_data in enumerate(data, start=last_row + 1):
    for col_index, value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index).value = value

# Save workbook
excel.save("C:\\Users\\MathijesiJohnbritto\\Documents\\read.xlsx")
